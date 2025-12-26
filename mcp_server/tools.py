# mcp_server/tools.py

from datetime import datetime
import logging

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from mcp_server.schema import DROPDOWNS
from utils.config_loader import load_config
from utils.backup import create_backup

# Load configuration
_config = load_config()
EXCEL_PATH = _config.get("excel_path", "data/employees_mcp.xlsx")
BACKUP_ENABLED = _config.get("processing", {}).get("backup_before_update", True)
BACKUP_DIR = _config.get("processing", {}).get("backup_directory", "backups")

# Setup logger
logger = logging.getLogger("excel_mcp.tools")

# Helper function to convert pandas/numpy types to JSON-serializable types
def convert_to_json_serializable(value):
    """Convert pandas/numpy types to Python native types for JSON serialization."""
    if pd.isna(value):
        return None
    elif isinstance(value, (np.integer, np.int64, np.int32, np.int8)):
        return int(value)
    elif isinstance(value, (np.floating, np.float64, np.float32, np.float16)):
        return float(value) if not pd.isna(value) else None
    elif isinstance(value, (np.bool_, bool)):
        return bool(value)
    elif hasattr(value, "item"):
        return value.item()
    elif isinstance(value, datetime):
        return value.isoformat()
    else:
        return value

def get_unprocessed_employees():
    """Fetch all unprocessed employee rows from Excel."""
    try:
        logger.info(f"Reading Excel file: {EXCEL_PATH}")
        df = pd.read_excel(EXCEL_PATH)
        
        # Handle case where Is_Processed column might not exist or have NaN values
        if "Is_Processed" not in df.columns:
            rows = df
            logger.warning("Is_Processed column not found, returning all rows")
        else:
            rows = df[df["Is_Processed"].fillna("No") != "Yes"]

        employees = []
        # Iterate through filtered rows and get their positional indices
        for idx, row in rows.iterrows():
            data = {}
            for k, v in row.to_dict().items():
                # Convert pandas/numpy types to Python native types for JSON serialization
                data[k] = convert_to_json_serializable(v)
            
            # Get the positional index (iloc position) in the original DataFrame
            # This works even after saving/reading because iloc uses position, not label
            try:
                positional_idx = df.index.get_loc(idx)
                # Convert to int if it's a numeric type
                if isinstance(positional_idx, (slice, np.ndarray)):
                    # Handle case where get_loc returns array (shouldn't happen with unique index)
                    positional_idx = int(positional_idx[0]) if hasattr(positional_idx, '__len__') else int(positional_idx)
                else:
                    positional_idx = int(positional_idx)
            except (KeyError, TypeError):
                # Fallback: use the index value directly if get_loc fails
                positional_idx = int(convert_to_json_serializable(idx)) if idx is not None else 0
            
            employees.append(
                {
                    "row_id": positional_idx,
                    "data": data,
                }
            )

        logger.info(f"Found {len(employees)} unprocessed employees")
        # return an object, not a bare list
        return {"employees": employees}
    except Exception as e:
        logger.error(f"Error fetching unprocessed employees: {e}", exc_info=True)
        raise


def get_all_employees():
    """Fetch ALL employee rows from Excel for comprehensive scanning."""
    try:
        logger.info(f"Reading Excel file: {EXCEL_PATH}")
        df = pd.read_excel(EXCEL_PATH)
        
        employees = []
        # Iterate through ALL rows
        for idx, row in df.iterrows():
            data = {}
            for k, v in row.to_dict().items():
                # Convert pandas/numpy types to Python native types for JSON serialization
                data[k] = convert_to_json_serializable(v)
            
            # Get the positional index
            try:
                positional_idx = df.index.get_loc(idx)
                if isinstance(positional_idx, (slice, np.ndarray)):
                    positional_idx = int(positional_idx[0]) if hasattr(positional_idx, '__len__') else int(positional_idx)
                else:
                    positional_idx = int(positional_idx)
            except (KeyError, TypeError):
                positional_idx = int(convert_to_json_serializable(idx)) if idx is not None else 0
            
            employees.append(
                {
                    "row_id": positional_idx,
                    "data": data,
                }
            )

        logger.info(f"Found {len(employees)} total employees for scanning")
        return {"employees": employees}
    except Exception as e:
        logger.error(f"Error fetching all employees: {e}", exc_info=True)
        raise
    except Exception as e:
        logger.error(f"Error fetching unprocessed employees: {e}", exc_info=True)
        raise



def update_employee_row(
    row_id: int,
    updates: dict,
    reason: str,
    confidence: float
):
    """Update employee row with AI decisions, with backup and validation."""
    try:
        logger.info(f"Updating row {row_id} with updates: {updates}")
        
        # Create backup before updating
        if BACKUP_ENABLED:
            backup_path = create_backup(EXCEL_PATH, BACKUP_DIR)
            if backup_path:
                logger.info(f"Created backup: {backup_path}")
            else:
                logger.warning("Backup creation failed, continuing without backup")
        
        df = pd.read_excel(EXCEL_PATH)
        
        # Validate row_id is a valid positional index
        if row_id < 0 or row_id >= len(df):
            error_msg = f"Invalid row_id: {row_id}. DataFrame has {len(df)} rows (0-{len(df)-1})"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Ensure required columns exist
        if "Is_Processed" not in df.columns:
            df["Is_Processed"] = "No"
        if "AI_Decision_Reason" not in df.columns:
            df["AI_Decision_Reason"] = ""
        if "Confidence_Score" not in df.columns:
            df["Confidence_Score"] = 0.0
        if "Last_Processed_On" not in df.columns:
            df["Last_Processed_On"] = None
        if "Processing_Version" not in df.columns:
            df["Processing_Version"] = None
        if "Rule_Applied" not in df.columns:
            df["Rule_Applied"] = None

        # 🔐 MCP VALIDATION - only validate non-empty values
        for col, val in updates.items():
            if col in DROPDOWNS and val and str(val).strip():  # Only validate if value is not empty
                if val not in DROPDOWNS[col]:
                    error_msg = f"Invalid value '{val}' for {col}. Allowed values: {DROPDOWNS[col]}"
                    logger.error(error_msg)
                    raise ValueError(error_msg)

        # Use iloc for positional indexing (since we save with index=False)
        # Only update columns that have non-empty values
        for col, val in updates.items():
            if val:  # Only update if value is not empty/None
                if col in df.columns:
                    df.iloc[row_id, df.columns.get_loc(col)] = val
                else:
                    logger.warning(f"Column '{col}' not found in DataFrame")

        df.iloc[row_id, df.columns.get_loc("Is_Processed")] = "Yes"
        df.iloc[row_id, df.columns.get_loc("AI_Decision_Reason")] = reason
        df.iloc[row_id, df.columns.get_loc("Confidence_Score")] = confidence
        df.iloc[row_id, df.columns.get_loc("Last_Processed_On")] = datetime.now()

        # Save while preserving dropdowns/data validation
        try:
            # Save updated DataFrame to Excel
            df.to_excel(EXCEL_PATH, index=False, engine='openpyxl')
            
            # Reapply data validation (dropdowns) after saving
            _reapply_data_validation(EXCEL_PATH, df.columns)
            
            logger.info(f"Successfully saved Excel file: {EXCEL_PATH}")
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}", exc_info=True)
            raise

        return {"status": "updated", "row_id": row_id, "updates_applied": updates}
    except Exception as e:
        logger.error(f"Error updating employee row {row_id}: {e}", exc_info=True)
        raise


def _reapply_data_validation(file_path, columns):
    """
    Reapply data validation (dropdowns) to columns after saving.
    This preserves dropdown functionality in Excel.
    """
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Get column indices for dropdown columns
        for col_name in columns:
            if col_name in DROPDOWNS:
                col_idx = list(columns).index(col_name) + 1  # +1 because Excel is 1-indexed
                col_letter = get_column_letter(col_idx)
                
                # Create data validation rule
                dropdown_values = DROPDOWNS[col_name]
                dv = DataValidation(type="list", formula1=f'"{",".join(dropdown_values)}"', allow_blank=True)
                
                # Apply to all data rows (skip header row 1)
                dv_range = f"{col_letter}2:{col_letter}{ws.max_row}"
                dv.add(dv_range)
                ws.add_data_validation(dv)
        
        wb.save(file_path)
        wb.close()
    except Exception as e:
        logger.warning(f"Could not reapply data validation: {e}")
        # Continue even if validation fails - at least data is saved


def update_experience_from_doj():
    """
    Recalculate Experience_Years for all employees based on their DOJ (Date of Joining).
    This should be run periodically (e.g., once a year) to update experience.
    """
    try:
        logger.info("Starting experience update from DOJ")
        
        # Create backup before updating
        if BACKUP_ENABLED:
            backup_path = create_backup(EXCEL_PATH, BACKUP_DIR)
            if backup_path:
                logger.info(f"Created backup: {backup_path}")
        
        df = pd.read_excel(EXCEL_PATH)
        
        if "DOJ" not in df.columns:
            error_msg = "DOJ column not found"
            logger.error(error_msg)
            return {"status": "error", "message": error_msg}
        if "Experience_Years" not in df.columns:
            df["Experience_Years"] = 0.0
        
        today = datetime.now().date()
        updated_count = 0
        doj_col_idx = df.columns.get_loc("DOJ")
        exp_col_idx = df.columns.get_loc("Experience_Years")
        
        for row_idx in range(len(df)):
            doj = df.iloc[row_idx, doj_col_idx]
            if pd.isna(doj):
                continue
            
            # Convert DOJ to date if it's not already
            if isinstance(doj, str):
                try:
                    doj = pd.to_datetime(doj).date()
                except (ValueError, TypeError):
                    logger.warning(f"Could not parse DOJ for row {row_idx}: {doj}")
                    continue
            elif hasattr(doj, 'date'):
                doj = doj.date()
            elif isinstance(doj, datetime):
                doj = doj.date()
            elif isinstance(doj, pd.Timestamp):
                doj = doj.date()
            
            # Calculate years of experience
            if isinstance(doj, type(today)):
                years_diff = today - doj
                experience_years = years_diff.days / 365.25  # Account for leap years
                df.iloc[row_idx, exp_col_idx] = round(experience_years, 1)
                updated_count += 1
        
        # Save the updated DataFrame while preserving dropdowns
        try:
            df.to_excel(EXCEL_PATH, index=False, engine='openpyxl')
            _reapply_data_validation(EXCEL_PATH, df.columns)
            logger.info(f"Updated experience for {updated_count} employees")
            return {"status": "success", "message": f"Updated experience for {updated_count} employees"}
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}", exc_info=True)
            return {"status": "error", "message": str(e)}
    except Exception as e:
        logger.error(f"Error updating experience: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}


def reset_processed_flag_for_reprocessing():
    """
    Reset Is_Processed flag to "No" for all employees to allow reprocessing.
    Useful when you want to reprocess all employees (e.g., after updating experience).
    """
    try:
        logger.info("Resetting Is_Processed flag for all employees")
        
        # Create backup before updating
        if BACKUP_ENABLED:
            backup_path = create_backup(EXCEL_PATH, BACKUP_DIR)
            if backup_path:
                logger.info(f"Created backup: {backup_path}")
        
        df = pd.read_excel(EXCEL_PATH)
        
        if "Is_Processed" not in df.columns:
            df["Is_Processed"] = "No"
        
        # Reset all rows to "No"
        df["Is_Processed"] = "No"
        reset_count = len(df)
        
        try:
            df.to_excel(EXCEL_PATH, index=False, engine='openpyxl')
            _reapply_data_validation(EXCEL_PATH, df.columns)
            logger.info(f"Reset Is_Processed flag for {reset_count} employees")
            return {"status": "success", "message": f"Reset Is_Processed flag for {reset_count} employees. They can now be reprocessed."}
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}", exc_info=True)
            return {"status": "error", "message": str(e)}
    except Exception as e:
        logger.error(f"Error resetting processed flag: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}
